# -*- coding: utf-8 -*-
"""
課程相關的 API views
包含課程搜尋、選課、退選、收藏等功能
"""
from django.contrib.auth.models import User
from django.views.decorators.csrf import csrf_exempt
from django.db.models import Q
from rest_framework.decorators import api_view
from rest_framework.response import Response
from .models import CourseOffering, Enrollment, FavoriteCourse, Profile
import openpyxl
from io import BytesIO


@api_view(['GET', 'POST'])
def search_courses(request):
    """搜尋課程"""
    try:
        # 支持 GET 和 POST 兩種方式取得參數
        if request.method == 'GET':
            keyword = request.GET.get('keyword', '').strip()
            department = request.GET.get('department', '').strip()
            course_type = request.GET.get('course_type', '').strip()
            semester = request.GET.get('semester', '').strip()
            grade_level = request.GET.get('grade_level', '').strip()
            academic_year = request.GET.get('academic_year', '114')
            
            # 取得複選參數（陣列）
            weekdays = request.GET.getlist('weekdays')  # ← 改這裡
            periods = request.GET.getlist('periods')    # ← 加這行
        else:  # POST
            keyword = request.data.get('keyword', '').strip()
            department = request.data.get('department', '').strip()
            course_type = request.data.get('course_type', '').strip()
            semester = request.data.get('semester', '').strip()
            grade_level = request.data.get('grade_level', '').strip()
            academic_year = request.data.get('academic_year', '114')
            
            # 取得複選參數（陣列）
            weekdays = request.data.get('weekdays', [])  # ← 改這裡
            periods = request.data.get('periods', [])    # ← 加這行
        
        print(f"搜尋條件: keyword={keyword}, department={department}, course_type={course_type}, semester={semester}, weekdays={weekdays}, periods={periods}, grade_level={grade_level}, academic_year={academic_year}")
        
        # 基本查詢：取得所有開課資料
        offerings = CourseOffering.objects.select_related(
            'course', 'department'
        ).prefetch_related(
            'offering_teachers__teacher__profile',
            'class_times'
        ).filter(
            academic_year=academic_year
        )
        
        # 應用篩選條件
        if semester:
            offerings = offerings.filter(semester=semester)
        
        if department:
            offerings = offerings.filter(department__name=department)
        
        if course_type:
            offerings = offerings.filter(course__course_type=course_type)
        
        if grade_level:
            offerings = offerings.filter(grade_level=int(grade_level))
        
        # 星期幾篩選（支援多選）
        if weekdays:
            offerings = offerings.filter(class_times__weekday__in=weekdays).distinct()
        
        # 節次篩選（支援多選）- 新增這段
        if periods:
            # 建立 Q 查詢來檢查時段是否有交集
            from django.db.models import Q
            period_query = Q()
            for period in periods:
                period_int = int(period)
                # 檢查該節次是否在 start_period 和 end_period 之間
                period_query |= Q(
                    class_times__start_period__lte=period_int,
                    class_times__end_period__gte=period_int
                )
            offerings = offerings.filter(period_query).distinct()
        
        # 關鍵字搜尋
        if keyword:
            offerings = offerings.filter(
                Q(course__course_name__icontains=keyword) |
                Q(course__course_code__icontains=keyword) |
                Q(offering_teachers__teacher__profile__real_name__icontains=keyword)
            ).distinct()
        
        # 組裝回傳資料
        courses_data = []
        for offering in offerings:
            # 取得所有上課時段
            class_times = offering.class_times.all()
            times_data = []
            for ct in class_times:
                times_data.append({
                    'weekday': ct.weekday,
                    'weekday_display': ct.get_weekday_display(),
                    'start_period': ct.start_period,
                    'end_period': ct.end_period,
                    'classroom': ct.classroom
                })
            
            # 取得所有教師
            teachers = offering.offering_teachers.all()
            teachers_data = []
            for ot in teachers:
                teacher_name = ot.teacher.profile.real_name if hasattr(ot.teacher, 'profile') else ot.teacher.username
                teachers_data.append({
                    'id': ot.teacher.id,
                    'name': teacher_name,
                    'role': ot.role,
                    'role_display': ot.get_role_display()
                })
            
            # 檢查是否已收藏
            is_favorited = False
            if request.user.is_authenticated:
                is_favorited = FavoriteCourse.objects.filter(
                    student=request.user,
                    offering=offering
                ).exists()
            
            courses_data.append({
                'id': offering.id,
                'course_code': offering.course.course_code,
                'course_name': offering.course.course_name,
                'course_type': offering.course.course_type,
                'course_type_display': offering.course.get_course_type_display(),
                'credits': offering.course.credits,
                'description': offering.course.description,
                'academic_year': offering.academic_year,
                'semester': offering.semester,
                'semester_display': offering.get_semester_display(),
                'department': offering.department.name,
                'grade_level': offering.grade_level,
                'teachers': teachers_data,
                'class_times': times_data,
                'max_students': offering.max_students,
                'current_students': offering.current_students,
                'status': offering.status,
                'status_display': offering.get_status_display(),
                'is_favorited': is_favorited,
            })
        
        print(f"找到 {len(courses_data)} 門課程")
        return Response(courses_data)
        
    except Exception as e:
        print(f"搜尋課程錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def enroll_course(request, course_id):
    """選課"""
    try:
        if not request.user.is_authenticated:
            return Response({'error': '請先登入'}, status=401)
        
        offering_id = course_id
        
        if not offering_id:
            return Response({'error': '缺少開課 ID'}, status=400)
        
        try:
            offering = CourseOffering.objects.get(id=offering_id)
        except CourseOffering.DoesNotExist:
            return Response({'error': '找不到該課程'}, status=404)
        
        # 檢查是否已選課
        if Enrollment.objects.filter(student=request.user, offering=offering, status='enrolled').exists():
            return Response({'error': '已經選過這門課'}, status=400)
        
        # 檢查是否額滿
        if offering.is_full():
            return Response({'error': '課程已額滿'}, status=400)
        
        # 建立選課記錄
        enrollment = Enrollment.objects.create(
            student=request.user,
            offering=offering,
            status='enrolled'
        )
        
        # 檢查時段衝突
        has_conflict, conflict_msg = enrollment.check_time_conflict()
        if has_conflict:
            enrollment.delete()
            return Response({'error': conflict_msg}, status=400)
        
        # 更新目前人數
        offering.current_students += 1
        if offering.current_students >= offering.max_students:
            offering.status = 'full'
        offering.save()
        
        print(f"{request.user.username} 選課成功: {offering.course.course_name}")
        return Response({'message': '選課成功'})
        
    except Exception as e:
        print(f"選課錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def drop_course(request, course_id):
    """退選"""
    try:
        if not request.user.is_authenticated:
            return Response({'error': '請先登入'}, status=401)
        
        offering_id = course_id
        
        if not offering_id:
            return Response({'error': '缺少開課 ID'}, status=400)
        
        try:
            enrollment = Enrollment.objects.get(
                student=request.user,
                offering_id=offering_id,
                status='enrolled'
            )
        except Enrollment.DoesNotExist:
            return Response({'error': '找不到選課記錄'}, status=404)
        
        offering = enrollment.offering
        
        # 更新狀態
        enrollment.status = 'dropped'
        enrollment.save()
        
        # 更新目前人數
        offering.current_students = max(0, offering.current_students - 1)
        if offering.status == 'full':
            offering.status = 'open'
        offering.save()
        
        print(f"{request.user.username} 退選成功: {offering.course.course_name}")
        return Response({'message': '退選成功'})
        
    except Exception as e:
        print(f"退選錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def get_enrolled_courses(request):
    """取得已選課程"""
    try:
        # 手動檢查登入狀態
        if not request.user.is_authenticated:
            print("使用者未登入")
            return Response([], status=200)
        
        # 獲取篩選參數
        academic_year = request.GET.get('academic_year', '114')
        semester = request.GET.get('semester', '1')
        
        print(f"取得 {request.user.username} 的選課記錄 (學年度: {academic_year}, 學期: {semester})")
        
        enrollments = Enrollment.objects.filter(
            student=request.user,
            status='enrolled',
            offering__academic_year=academic_year,
            offering__semester=semester
        ).select_related(
            'offering__course',
            'offering__department'
        ).prefetch_related(
            'offering__offering_teachers__teacher__profile',
            'offering__class_times'
        )
        
        courses_data = []
        for enrollment in enrollments:
            offering = enrollment.offering
            
            # 取得上課時段
            class_times = offering.class_times.all()
            times_data = []
            for ct in class_times:
                times_data.append({
                    'weekday': ct.weekday,
                    'weekday_display': ct.get_weekday_display(),
                    'start_period': ct.start_period,
                    'end_period': ct.end_period,
                    'classroom': ct.classroom
                })
            
            # 取得所有教師
            teachers = offering.offering_teachers.all()
            teachers_data = []
            for ot in teachers:
                teacher_name = ot.teacher.profile.real_name if hasattr(ot.teacher, 'profile') else ot.teacher.username
                teachers_data.append({
                    'id': ot.teacher.id,
                    'name': teacher_name,
                    'role': ot.role,
                })
            
            # 取得主教師
            main_teacher = offering.offering_teachers.filter(role='main').first()
            teacher_name = main_teacher.teacher.profile.real_name if main_teacher and hasattr(main_teacher.teacher, 'profile') else '未設定'
            
            courses_data.append({
                'id': offering.id,
                'course_code': offering.course.course_code,
                'course_name': offering.course.course_name,
                'course_type': offering.course.course_type,
                'course_type_display': offering.course.get_course_type_display(),
                'credits': offering.course.credits,
                'teacher_name': teacher_name,
                'teachers': teachers_data,
                'class_times': times_data,
                'enrolled_at': enrollment.enrolled_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        print(f"找到 {len(courses_data)} 門已選課程")
        return Response(courses_data)
        
    except Exception as e:
        print(f"錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def toggle_favorite(request, course_id):
    """收藏/取消收藏課程"""
    try:
        if not request.user.is_authenticated:
            return Response({'error': '請先登入'}, status=401)
        
        offering_id = course_id
        
        if not offering_id:
            return Response({'error': '缺少開課 ID'}, status=400)
        
        try:
            offering = CourseOffering.objects.get(id=offering_id)
        except CourseOffering.DoesNotExist:
            return Response({'error': '找不到該課程'}, status=404)
        
        # 檢查是否已收藏
        favorite = FavoriteCourse.objects.filter(
            student=request.user,
            offering=offering
        ).first()
        
        if favorite:
            # 已收藏，取消收藏
            favorite.delete()
            return Response({'message': '已取消收藏', 'is_favorited': False})
        else:
            # 未收藏，新增收藏
            FavoriteCourse.objects.create(
                student=request.user,
                offering=offering
            )
            return Response({'message': '已加入收藏', 'is_favorited': True})
        
    except Exception as e:
        print(f"收藏錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def get_favorite_courses(request):
    """取得收藏的課程"""
    try:
        if not request.user.is_authenticated:
            return Response({'error': '請先登入'}, status=401)
        
        favorites = FavoriteCourse.objects.filter(
            student=request.user
        ).select_related(
            'offering__course',
            'offering__department'
        ).prefetch_related(
            'offering__offering_teachers__teacher__profile',
            'offering__class_times'
        )
        
        courses_data = []
        for favorite in favorites:
            offering = favorite.offering
            
            # 取得上課時段
            class_times = offering.class_times.all()
            times_data = []
            for ct in class_times:
                times_data.append({
                    'weekday': ct.weekday,
                    'weekday_display': ct.get_weekday_display(),
                    'start_period': ct.start_period,
                    'end_period': ct.end_period,
                    'classroom': ct.classroom
                })
            
            # 取得所有教師
            teachers = offering.offering_teachers.all()
            teachers_data = []
            for ot in teachers:
                teacher_name = ot.teacher.profile.real_name if hasattr(ot.teacher, 'profile') else ot.teacher.username
                teachers_data.append({
                    'id': ot.teacher.id,
                    'name': teacher_name,
                    'role': ot.role,
                    'role_display': ot.get_role_display()
                })
            
            # 取得主教師
            main_teacher = offering.offering_teachers.filter(role='main').first()
            teacher_name = main_teacher.teacher.profile.real_name if main_teacher and hasattr(main_teacher.teacher, 'profile') else '未設定'
            
            courses_data.append({
                'id': offering.id,
                'course_code': offering.course.course_code,
                'course_name': offering.course.course_name,
                'course_type': offering.course.course_type,
                'course_type_display': offering.course.get_course_type_display(),
                'credits': offering.course.credits,
                'description': offering.course.description,
                'academic_year': offering.academic_year,
                'semester': offering.semester,
                'semester_display': offering.get_semester_display(),
                'department': offering.department.name,
                'grade_level': offering.grade_level,
                'teacher_name': teacher_name,
                'teachers': teachers_data,
                'class_times': times_data,
                'max_students': offering.max_students,
                'current_students': offering.current_students,
                'status': offering.status,
                'status_display': offering.get_status_display(),
                'is_favorited': True,  # 收藏列表中的課程當然都是已收藏
                'favorited_at': favorite.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            })
        
        return Response(courses_data)
        
    except Exception as e:
        print(f"錯誤: {str(e)}")
        return Response({'error': str(e)}, status=500)


@api_view(['POST'])
def import_courses_excel(request):
    """從 Excel 匯入課程"""
    try:
        if 'file' not in request.FILES:
            return Response({'error': '沒有上傳檔案'}, status=400)
        
        excel_file = request.FILES['file']
        
        # 讀取 Excel
        wb = openpyxl.load_workbook(BytesIO(excel_file.read()))
        ws = wb.active
        
        # 取得欄位數量以判斷格式
        first_row = next(ws.iter_rows(min_row=1, max_row=1))
        column_count = len([cell for cell in first_row if cell.value is not None])
        
        print(f"偵測到 {column_count} 欄")
        
        success_count = 0
        error_count = 0
        errors = []
        
        # 根據欄位數量選擇對應的處理方式
        if column_count == 15:
            # 15 欄格式
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:  # 跳過空行
                        continue
                    
                    # 處理邏輯
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"第 {idx} 列: {str(e)}")
        
        elif column_count == 16:
            # 16 欄格式（含系所）
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    if not row[0]:
                        continue
                    
                    # 處理邏輯
                    success_count += 1
                    
                except Exception as e:
                    error_count += 1
                    errors.append(f"第 {idx} 列: {str(e)}")
        
        else:
            return Response({'error': f'不支援的格式（{column_count} 欄）'}, status=400)
        
        result = {
            'message': f'匯入完成: 成功 {success_count} 筆，失敗 {error_count} 筆',
            'success_count': success_count,
            'error_count': error_count,
        }
        
        if errors:
            result['errors'] = errors[:10]  # 只回傳前 10 個錯誤
        
        return Response(result)
        
    except Exception as e:
        print(f"匯入錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def get_filter_options(request):
    """取得篩選選項（系所、學期等）"""
    try:
        from .models import Department
        
        # 取得所有系所
        departments = Department.objects.all().values_list('name', flat=True).distinct()
        
        # 取得所有學年度
        academic_years = CourseOffering.objects.values_list('academic_year', flat=True).distinct().order_by('-academic_year')
        
        # 學期選項
        semesters = [
            {'value': '1', 'label': '上學期'},
            {'value': '2', 'label': '下學期'},
        ]
        
        # 課程類別選項
        course_types = [
            {'value': 'required', 'label': '必修'},
            {'value': 'elective', 'label': '選修'},
            {'value': 'general_required', 'label': '通識(必修)'},
            {'value': 'general_elective', 'label': '通識(選修)'},
        ]
        
        # 星期選項
        weekdays = [
            {'value': '1', 'label': '星期一'},
            {'value': '2', 'label': '星期二'},
            {'value': '3', 'label': '星期三'},
            {'value': '4', 'label': '星期四'},
            {'value': '5', 'label': '星期五'},
            {'value': '6', 'label': '星期六'},
            {'value': '7', 'label': '星期日'},
        ]
        
        # 年級選項
        grades = [
            {'value': '1', 'label': '一年級'},
            {'value': '2', 'label': '二年級'},
            {'value': '3', 'label': '三年級'},
            {'value': '4', 'label': '四年級'},
        ]
        
        return Response({
            'departments': list(departments),
            'academic_years': list(academic_years),
            'semesters': semesters,
            'course_types': course_types,
            'weekdays': weekdays,
            'grades': grades,
        })
        
    except Exception as e:
        print(f"取得篩選選項錯誤: {str(e)}")
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def get_course_detail(request, course_id):
    """取得單一課程詳細資料（用於編輯）"""
    try:
        offering = CourseOffering.objects.select_related(
            'course', 'department'
        ).prefetch_related(
            'offering_teachers__teacher__profile',
            'class_times'
        ).get(id=course_id)
        
        # 取得上課時段（取第一個時段）
        class_time = offering.class_times.first()
        
        # 取得主教師
        main_teacher = offering.offering_teachers.filter(role='main').first()
        
        course_data = {
            'id': offering.id,
            'course_code': offering.course.course_code,
            'course_name': offering.course.course_name,
            'course_type': offering.course.course_type,
            'description': offering.course.description or '',
            'credits': offering.course.credits,
            'hours': offering.course.credits,  # 假設小時數等於學分數
            'academic_year': offering.academic_year,
            'semester': offering.semester,
            'department': offering.department.name,
            'grade_level': str(offering.grade_level),
            'teacher_id': str(main_teacher.teacher.id) if main_teacher else '',
            'classroom': class_time.classroom if class_time else '',
            'weekday': class_time.weekday if class_time else '1',
            'start_period': str(class_time.start_period) if class_time else '1',
            'end_period': str(class_time.end_period) if class_time else '2',
            'max_students': offering.max_students,
        }
        
        return Response(course_data)
        
    except CourseOffering.DoesNotExist:
        return Response({'error': '找不到該課程'}, status=404)
    except Exception as e:
        print(f"取得課程詳情錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['PUT'])
def update_course(request, course_id):
    """更新課程資料"""
    try:
        # 取得開課資料
        offering = CourseOffering.objects.select_related('course').get(id=course_id)
        course = offering.course
        
        # 更新課程基本資料
        course.course_code = request.data.get('course_code', course.course_code)
        course.course_name = request.data.get('course_name', course.course_name)
        course.course_type = request.data.get('course_type', course.course_type)
        course.description = request.data.get('description', course.description)
        course.credits = request.data.get('credits', course.credits)
        course.save()
        
        # 更新開課資料
        offering.academic_year = request.data.get('academic_year', offering.academic_year)
        offering.semester = request.data.get('semester', offering.semester)
        offering.grade_level = request.data.get('grade_level', offering.grade_level)
        offering.max_students = request.data.get('max_students', offering.max_students)
        
        # 更新系所
        from .models import Department
        department_name = request.data.get('department')
        if department_name:
            department, _ = Department.objects.get_or_create(name=department_name)
            offering.department = department
        
        offering.save()
        
        # 更新教師
        teacher_id = request.data.get('teacher_id')
        if teacher_id:
            from .models import OfferingTeacher
            # 刪除舊的教師關係
            offering.offering_teachers.all().delete()
            # 新增新的教師關係
            teacher = User.objects.get(id=teacher_id)
            OfferingTeacher.objects.create(
                offering=offering,
                teacher=teacher,
                role='main'
            )
        
        # 更新上課時段
        classroom = request.data.get('classroom')
        weekday = request.data.get('weekday')
        start_period = request.data.get('start_period')
        end_period = request.data.get('end_period')
        
        if all([classroom, weekday, start_period, end_period]):
            from .models import ClassTime
            # 刪除舊的時段
            offering.class_times.all().delete()
            # 新增新的時段
            ClassTime.objects.create(
                offering=offering,
                weekday=weekday,
                start_period=start_period,
                end_period=end_period,
                classroom=classroom
            )
        
        print(f"課程更新成功: {course.course_name}")
        return Response({'message': '課程更新成功'})
        
    except CourseOffering.DoesNotExist:
        return Response({'error': '找不到該課程'}, status=404)
    except User.DoesNotExist:
        return Response({'error': '找不到該教師'}, status=404)
    except Exception as e:
        print(f"更新課程錯誤: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)


@api_view(['GET'])
def my_teaching_courses(request):
    """取得教師自己的授課列表（包含主開課和協同）"""
    try:
        # 1. 檢查登入
        if not request.user.is_authenticated:
            return Response({'error': '請先登入'}, status=401)
            
        print(f"查詢教師課程: {request.user.username}")
        
        # 2. 查詢該教師的所有開課（無論是主開課還是協同）
        # 使用 offering_teachers__teacher 關聯查詢
        offerings = CourseOffering.objects.filter(
            offering_teachers__teacher=request.user
        ).select_related(
            'course', 'department'
        ).prefetch_related(
            'offering_teachers__teacher__profile',
            'class_times'
        ).distinct().order_by('-academic_year', '-semester', 'course__course_code')
        
        # 3. 整理回傳資料
        courses_data = []
        for offering in offerings:
            # 取得該教師在這門課的角色
            my_role_rel = offering.offering_teachers.filter(teacher=request.user).first()
            my_role = my_role_rel.get_role_display() if my_role_rel else '未知'
            
            # 取得上課時段
            class_times = offering.class_times.all()
            times_display = []
            for ct in class_times:
                times_display.append(f"{ct.get_weekday_display()} 第{ct.start_period}-{ct.end_period}節 ({ct.classroom})")
            
            # 取得所有教師名稱
            all_teachers = offering.offering_teachers.all()
            teacher_names = []
            for ot in all_teachers:
                name = ot.teacher.profile.real_name if hasattr(ot.teacher, 'profile') else ot.teacher.username
                if ot.role == 'main':
                    teacher_names.append(f"{name}(主)")
                else:
                    teacher_names.append(name)
            
            courses_data.append({
                'id': offering.id,
                'course_code': offering.course.course_code,
                'course_name': offering.course.course_name,
                'course_type': offering.course.get_course_type_display(),
                'credits': offering.course.credits,
                'academic_year': offering.academic_year,
                'semester': offering.get_semester_display(),
                'department': offering.department.name,
                'my_role': my_role,
                'time_info': '；'.join(times_display) if times_display else '未設定',
                'teacher_names': '、'.join(teacher_names),
                'student_count': f"{offering.current_students} / {offering.max_students}",
                'status': offering.get_status_display()
            })
            
        print(f"找到 {len(courses_data)} 門授課")
        return Response(courses_data)

    except Exception as e:
        print(f"取得授課列表失敗: {str(e)}")
        import traceback
        traceback.print_exc()
        return Response({'error': str(e)}, status=500)